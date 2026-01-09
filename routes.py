from flask import render_template, request, redirect, url_for, flash, send_file
from models import (Employee, Attendance, Advance, Site, SiteWorker, 
                   MaterialCategory, SiteMaterial, MaterialPayment, SiteExpense)
from datetime import datetime, timedelta
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from database import get_db
from psycopg2.extras import RealDictCursor

def init_routes(app):
    
    # ========================================
    # DASHBOARD
    # ========================================
    
    @app.route('/')
    def dashboard():
        total_employees = Employee.count()
        return render_template('dashboard.html', total_employees=total_employees)
    
    # ========================================
    # EMPLOYEE ROUTES
    # ========================================
    
    @app.route('/employees')
    def employees():
        employees = Employee.get_all()
        return render_template('employees.html', employees=employees)
    
    @app.route('/add_employee', methods=['GET', 'POST'])
    def add_employee():
        if request.method == 'POST':
            name = request.form['name']
            role = request.form['role']
            daily_salary = float(request.form['daily_salary'])
            Employee.create(name, role, daily_salary)
            flash('Employee added successfully!', 'success')
            return redirect(url_for('employees'))
        return render_template('add_employee.html')
    
    @app.route('/edit_employee/<int:emp_id>', methods=['GET', 'POST'])
    def edit_employee(emp_id):
        if request.method == 'POST':
            name = request.form['name']
            role = request.form['role']
            daily_salary = float(request.form['daily_salary'])
            Employee.update(emp_id, name, role, daily_salary)
            flash('Employee updated successfully!', 'success')
            return redirect(url_for('employees'))
        employee = Employee.get_by_id(emp_id)
        return render_template('edit_employee.html', employee=employee)
    
    @app.route('/delete_employee/<int:emp_id>')
    def delete_employee(emp_id):
        Employee.delete(emp_id)
        flash('Employee deleted successfully!', 'success')
        return redirect(url_for('employees'))
    
    # ========================================
    # ATTENDANCE ROUTES
    # ========================================
    
    @app.route('/attendance', methods=['GET', 'POST'])
    def attendance():
        if request.method == 'POST':
            date = request.form['date']
            employees = Employee.get_all()
            for emp in employees:
                status = request.form.get(f'status_{emp["id"]}')
                if status:
                    Attendance.mark(emp['id'], date, status)
            flash('Attendance marked successfully!', 'success')
            return redirect(url_for('attendance'))
        
        date = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
        employees = Employee.get_all()
        attendance_records = Attendance.get_by_date(date)
        
        attendance_dict = {rec['employee_id']: rec['status'] for rec in attendance_records}
        
        return render_template('attendance.html', 
                             employees=employees, 
                             date=date,
                             attendance_dict=attendance_dict)
    
    # ========================================
    # ADVANCE ROUTES
    # ========================================
    
    @app.route('/advance', methods=['GET', 'POST'])
    def advance():
        if request.method == 'POST':
            employee_id = int(request.form['employee_id'])
            date = request.form['date']
            amount = float(request.form['amount'])
            reason = request.form.get('reason', '')
            Advance.create(employee_id, date, amount, reason)
            flash('Advance payment recorded successfully!', 'success')
            return redirect(url_for('advance'))
        
        employees = Employee.get_all()
        advances = Advance.get_all()
        return render_template('advance.html', employees=employees, advances=advances)
    
    # ========================================
    # WEEKLY PAYROLL ROUTES
    # ========================================
    
    @app.route('/weekly_payroll', methods=['GET', 'POST'])
    def weekly_payroll():
        if request.method == 'POST':
            start_date = request.form['start_date']
            end_date = request.form['end_date']
        else:
            today = datetime.now()
            start_date = request.args.get('start_date', 
                                         (today - timedelta(days=today.weekday())).strftime('%Y-%m-%d'))
            end_date = request.args.get('end_date', 
                                       (today + timedelta(days=6-today.weekday())).strftime('%Y-%m-%d'))
        
        employees = Employee.get_all()
        payroll_data = []
        total_payroll = 0
        
        for emp in employees:
            present_days = Attendance.get_week_attendance(emp['id'], start_date, end_date)
            gross_salary = present_days * emp['daily_salary']
            total_advance = Advance.get_week_advance(emp['id'], start_date, end_date)
            net_salary = float(gross_salary) - float(total_advance)
            total_payroll += net_salary
            
            payroll_data.append({
                'employee': emp,
                'present_days': present_days,
                'gross_salary': float(gross_salary),
                'total_advance': float(total_advance),
                'net_salary': net_salary
            })
        
        return render_template('weekly_payroll.html', 
                             payroll_data=payroll_data,
                             start_date=start_date,
                             end_date=end_date,
                             total_payroll=total_payroll)
    
    @app.route('/export_weekly_excel')
    def export_weekly_excel():
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Weekly Payroll"
        
        ws.merge_cells('A1:H1')
        ws['A1'] = 'SAVUNADRY CONSTRUCTION'
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells('A2:H2')
        ws['A2'] = f'Weekly Payroll Report'
        ws['A2'].font = Font(size=12, bold=True)
        ws['A2'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells('A3:H3')
        ws['A3'] = f'Period: {start_date} to {end_date}'
        ws['A3'].font = Font(size=10)
        ws['A3'].alignment = Alignment(horizontal='center')
        
        headers = ['Employee ID', 'Name', 'Role', 'Present Days', 'Daily Salary', 'Gross Salary', 'Advance', 'Net Salary']
        ws.append([])
        ws.append(headers)
        
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        for cell in ws[5]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        employees = Employee.get_all()
        total_payroll = 0
        
        for emp in employees:
            present_days = Attendance.get_week_attendance(emp['id'], start_date, end_date)
            gross_salary = present_days * float(emp['daily_salary'])
            total_advance = float(Advance.get_week_advance(emp['id'], start_date, end_date))
            net_salary = gross_salary - total_advance
            total_payroll += net_salary
            
            ws.append([
                emp['id'], emp['name'], emp['role'], present_days,
                float(emp['daily_salary']), gross_salary, total_advance, net_salary
            ])
        
        ws.append([])
        total_row = ws.max_row
        ws[f'A{total_row}'] = 'TOTAL'
        ws[f'A{total_row}'].font = Font(bold=True)
        ws[f'H{total_row}'] = total_payroll
        ws[f'H{total_row}'].font = Font(bold=True)
        
        for row in ws.iter_rows(min_row=6, max_row=ws.max_row, min_col=5, max_col=8):
            for cell in row:
                cell.number_format = '₹#,##0.00'
        
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws.column_dimensions[col].width = 15
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(output,
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        as_attachment=True,
                        download_name=f'weekly_payroll_{start_date}_to_{end_date}.xlsx')
    
    @app.route('/export_weekly_pdf')
    def export_weekly_pdf():
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=landscape(A4))
        elements = []
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'],
                                     fontSize=18, spaceAfter=10, alignment=TA_CENTER)
        
        elements.append(Paragraph('SAVUNADRY CONSTRUCTION', title_style))
        elements.append(Paragraph(f'Weekly Payroll Report', styles['Heading2']))
        elements.append(Paragraph(f'Period: {start_date} to {end_date}', styles['Normal']))
        elements.append(Spacer(1, 20))
        
        employees = Employee.get_all()
        data = [['ID', 'Name', 'Role', 'Days', 'Daily Salary', 'Gross', 'Advance', 'Net Salary']]
        total_payroll = 0
        
        for emp in employees:
            present_days = Attendance.get_week_attendance(emp['id'], start_date, end_date)
            gross_salary = present_days * float(emp['daily_salary'])
            total_advance = float(Advance.get_week_advance(emp['id'], start_date, end_date))
            net_salary = gross_salary - total_advance
            total_payroll += net_salary
            
            data.append([str(emp['id']), emp['name'], emp['role'], str(present_days),
                        f"₹{emp['daily_salary']:.2f}", f"₹{gross_salary:.2f}",
                        f"₹{total_advance:.2f}", f"₹{net_salary:.2f}"])
        
        data.append(['', '', '', '', '', '', 'TOTAL:', f"₹{total_payroll:.2f}"])
        
        table = Table(data, colWidths=[0.6*inch, 1.5*inch, 1.2*inch, 0.8*inch, 1*inch, 1*inch, 1*inch, 1.2*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E7E6E6')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(table)
        doc.build(elements)
        output.seek(0)
        
        return send_file(output, mimetype='application/pdf', as_attachment=True,
                        download_name=f'weekly_payroll_{start_date}_to_{end_date}.pdf')
    
    # ========================================
    # MONTHLY REPORT ROUTES
    # ========================================
    
    @app.route('/monthly_report', methods=['GET', 'POST'])
    def monthly_report():
        if request.method == 'POST':
            month = request.form['month']
            year = request.form['year']
        else:
            today = datetime.now()
            month = request.args.get('month', f'{today.month:02d}')
            year = request.args.get('year', str(today.year))
        
        employees = Employee.get_all()
        report_data = []
        total_salary = 0
        total_advance = 0
        total_net = 0
        
        for emp in employees:
            present_days = Attendance.get_month_attendance(emp['id'], month, year)
            gross_salary = present_days * emp['daily_salary']
            advance_amount = Advance.get_month_advance(emp['id'], month, year)
            net_amount = float(gross_salary) - float(advance_amount)
            
            total_salary += float(gross_salary)
            total_advance += float(advance_amount)
            total_net += net_amount
            
            report_data.append({
                'employee': emp,
                'present_days': present_days,
                'gross_salary': float(gross_salary),
                'total_advance': float(advance_amount),
                'net_amount': net_amount
            })
        
        return render_template('monthly_report.html', 
                             report_data=report_data, month=month, year=year,
                             total_salary=total_salary, total_advance=total_advance,
                             total_net=total_net)
    
    @app.route('/export_monthly_excel')
    def export_monthly_excel():
        month = request.args.get('month')
        year = request.args.get('year')
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Monthly Report"
        
        ws.merge_cells('A1:G1')
        ws['A1'] = 'SAVUNADRY CONSTRUCTION'
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        month_names = ['', 'January', 'February', 'March', 'April', 'May', 'June',
                      'July', 'August', 'September', 'October', 'November', 'December']
        ws.merge_cells('A2:G2')
        ws['A2'] = f'Monthly Report - {month_names[int(month)]} {year}'
        ws['A2'].font = Font(size=12, bold=True)
        ws['A2'].alignment = Alignment(horizontal='center')
        
        headers = ['Employee ID', 'Name', 'Role', 'Days Worked', 'Gross Salary', 'Advance', 'Net Paid']
        ws.append([])
        ws.append(headers)
        
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        for cell in ws[4]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        employees = Employee.get_all()
        total_salary = 0
        total_advance = 0
        total_net = 0
        
        for emp in employees:
            present_days = Attendance.get_month_attendance(emp['id'], month, year)
            gross_salary = present_days * float(emp['daily_salary'])
            advance_amount = float(Advance.get_month_advance(emp['id'], month, year))
            net_amount = gross_salary - advance_amount
            
            total_salary += gross_salary
            total_advance += advance_amount
            total_net += net_amount
            
            ws.append([emp['id'], emp['name'], emp['role'], present_days,
                      gross_salary, advance_amount, net_amount])
        
        ws.append([])
        total_row = ws.max_row
        ws[f'A{total_row}'] = 'TOTAL'
        ws[f'E{total_row}'] = total_salary
        ws[f'F{total_row}'] = total_advance
        ws[f'G{total_row}'] = total_net
        for col in ['A', 'E', 'F', 'G']:
            ws[f'{col}{total_row}'].font = Font(bold=True)
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(output,
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        as_attachment=True,
                        download_name=f'monthly_report_{month}_{year}.xlsx')
    
    @app.route('/export_monthly_pdf')
    def export_monthly_pdf():
        month = request.args.get('month')
        year = request.args.get('year')
        
        month_names = ['', 'January', 'February', 'March', 'April', 'May', 'June',
                      'July', 'August', 'September', 'October', 'November', 'December']
        
        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=landscape(A4))
        elements = []
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'],
                                     fontSize=18, spaceAfter=10, alignment=TA_CENTER)
        
        elements.append(Paragraph('SAVUNADRY CONSTRUCTION', title_style))
        elements.append(Paragraph(f'Monthly Report - {month_names[int(month)]} {year}', styles['Heading2']))
        elements.append(Spacer(1, 20))
        
        employees = Employee.get_all()
        data = [['ID', 'Name', 'Role', 'Days', 'Gross', 'Advance', 'Net']]
        total_salary = 0
        total_advance = 0
        total_net = 0
        
        for emp in employees:
            present_days = Attendance.get_month_attendance(emp['id'], month, year)
            gross_salary = present_days * float(emp['daily_salary'])
            advance_amount = float(Advance.get_month_advance(emp['id'], month, year))
            net_amount = gross_salary - advance_amount
            
            total_salary += gross_salary
            total_advance += advance_amount
            total_net += net_amount
            
            data.append([str(emp['id']), emp['name'], emp['role'], str(present_days),
                        f"₹{gross_salary:.2f}", f"₹{advance_amount:.2f}", f"₹{net_amount:.2f}"])
        
        data.append(['', '', '', 'TOTAL:', f"₹{total_salary:.2f}", f"₹{total_advance:.2f}", f"₹{total_net:.2f}"])
        
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(table)
        doc.build(elements)
        output.seek(0)
        
        return send_file(output, mimetype='application/pdf', as_attachment=True,
                        download_name=f'monthly_report_{month}_{year}.pdf')
    
    # ========================================
    # SITE MANAGEMENT ROUTES
    # ========================================
    
    @app.route('/sites')
    def sites():
        sites = Site.get_all()
        return render_template('sites.html', sites=sites)
    
    @app.route('/add_site', methods=['GET', 'POST'])
    def add_site():
        if request.method == 'POST':
            site_name = request.form['site_name']
            location = request.form['location']
            client_name = request.form.get('client_name', '')
            start_date = request.form.get('start_date') or None
            end_date = request.form.get('end_date') or None
            total_budget = float(request.form.get('total_budget', 0))
            notes = request.form.get('notes', '')
            
            site_id = Site.create(site_name, location, client_name, start_date, end_date, total_budget, notes)
            if site_id:
                flash('Site added successfully!', 'success')
                return redirect(url_for('site_detail', site_id=site_id))
            else:
                flash('Error adding site!', 'error')
        
        return render_template('add_site.html')
    
    @app.route('/edit_site/<int:site_id>', methods=['GET', 'POST'])
    def edit_site(site_id):
        if request.method == 'POST':
            site_name = request.form['site_name']
            location = request.form['location']
            client_name = request.form.get('client_name', '')
            start_date = request.form.get('start_date') or None
            end_date = request.form.get('end_date') or None
            status = request.form.get('status', 'Active')
            total_budget = float(request.form.get('total_budget', 0))
            notes = request.form.get('notes', '')
            
            if Site.update(site_id, site_name, location, client_name, start_date, end_date, status, total_budget, notes):
                flash('Site updated successfully!', 'success')
                return redirect(url_for('site_detail', site_id=site_id))
            else:
                flash('Error updating site!', 'error')
        
        site = Site.get_by_id(site_id)
        return render_template('edit_site.html', site=site)
    
    @app.route('/delete_site/<int:site_id>')
    def delete_site(site_id):
        if Site.delete(site_id):
            flash('Site deleted successfully!', 'success')
        else:
            flash('Error deleting site!', 'error')
        return redirect(url_for('sites'))
    
    @app.route('/site_detail/<int:site_id>')
    def site_detail(site_id):
        site = Site.get_summary(site_id)
        workers = SiteWorker.get_site_workers(site_id)
        materials = SiteMaterial.get_site_materials(site_id)
        expenses = SiteExpense.get_site_expenses(site_id)
        
        material_summary = {}
        for material in materials:
            category = material['category_name']
            if category not in material_summary:
                material_summary[category] = {'total_cost': 0, 'total_paid': 0, 'total_balance': 0, 'count': 0}
            material_summary[category]['total_cost'] += float(material['total_cost'])
            material_summary[category]['total_paid'] += float(material['amount_paid'])
            material_summary[category]['total_balance'] += float(material['amount_balance'])
            material_summary[category]['count'] += 1
        
        return render_template('site_detail.html', site=site, workers=workers,
                             materials=materials, expenses=expenses,
                             material_summary=material_summary)
    
    @app.route('/assign_worker/<int:site_id>', methods=['GET', 'POST'])
    def assign_worker(site_id):
        if request.method == 'POST':
            employee_id = int(request.form['employee_id'])
            assigned_date = request.form['assigned_date']
            role_at_site = request.form.get('role_at_site', '')
            
            if SiteWorker.assign_worker(site_id, employee_id, assigned_date, role_at_site):
                flash('Worker assigned successfully!', 'success')
            else:
                flash('Error assigning worker!', 'error')
            
            return redirect(url_for('site_detail', site_id=site_id))
        
        site = Site.get_by_id(site_id)
        employees = Employee.get_all()
        current_workers = SiteWorker.get_site_workers(site_id)
        current_worker_ids = [w['employee_id'] for w in current_workers]
        available_employees = [e for e in employees if e['id'] not in current_worker_ids]
        
        return render_template('assign_worker.html', site=site, employees=available_employees)
    
    @app.route('/remove_worker/<int:site_worker_id>')
    def remove_worker(site_worker_id):
        removed_date = datetime.now().strftime('%Y-%m-%d')
        if SiteWorker.remove_worker(site_worker_id, removed_date):
            flash('Worker removed from site!', 'success')
        else:
            flash('Error removing worker!', 'error')
        return redirect(request.referrer or url_for('sites'))
    
    @app.route('/add_material/<int:site_id>', methods=['GET', 'POST'])
    def add_material(site_id):
        if request.method == 'POST':
            material_category_id = int(request.form['material_category_id'])
            material_name = request.form['material_name']
            quantity = float(request.form['quantity'])
            unit = request.form['unit']
            rate_per_unit = float(request.form['rate_per_unit'])
            supplier_name = request.form.get('supplier_name', '')
            sent_date = request.form['sent_date']
            bill_number = request.form.get('bill_number', '')
            notes = request.form.get('notes', '')
            
            if SiteMaterial.create(site_id, material_category_id, material_name, quantity, unit,
                                  rate_per_unit, supplier_name, sent_date, bill_number, notes):
                flash('Material added successfully!', 'success')
                return redirect(url_for('site_detail', site_id=site_id))
            else:
                flash('Error adding material!', 'error')
        
        site = Site.get_by_id(site_id)
        categories = MaterialCategory.get_all()
        return render_template('add_material.html', site=site, categories=categories)
    
    @app.route('/material_detail/<int:material_id>')
    def material_detail(material_id):
        material = SiteMaterial.get_by_id(material_id)
        payments = MaterialPayment.get_material_payments(material_id)
        return render_template('material_detail.html', material=material, payments=payments)
    
    @app.route('/add_payment/<int:material_id>', methods=['GET', 'POST'])
    def add_payment(material_id):
        if request.method == 'POST':
            payment_date = request.form['payment_date']
            amount = float(request.form['amount'])
            payment_mode = request.form.get('payment_mode', '')
            reference_number = request.form.get('reference_number', '')
            notes = request.form.get('notes', '')
            
            if MaterialPayment.create(material_id, payment_date, amount, payment_mode, reference_number, notes):
                flash('Payment recorded successfully!', 'success')
                return redirect(url_for('material_detail', material_id=material_id))
            else:
                flash('Error recording payment!', 'error')
        
        material = SiteMaterial.get_by_id(material_id)
        return render_template('add_payment.html', material=material)
    
    @app.route('/pending_payments')
    def pending_payments():
        materials = SiteMaterial.get_pending_payments()
        return render_template('pending_payments.html', materials=materials)
    
    @app.route('/add_expense/<int:site_id>', methods=['GET', 'POST'])
    def add_expense(site_id):
        if request.method == 'POST':
            expense_date = request.form['expense_date']
            expense_type = request.form['expense_type']
            description = request.form.get('description', '')
            amount = float(request.form['amount'])
            paid_to = request.form.get('paid_to', '')
            payment_mode = request.form.get('payment_mode', '')
            
            if SiteExpense.create(site_id, expense_date, expense_type, description, amount, paid_to, payment_mode):
                flash('Expense recorded successfully!', 'success')
                return redirect(url_for('site_detail', site_id=site_id))
            else:
                flash('Error recording expense!', 'error')
        
        site = Site.get_by_id(site_id)
        return render_template('add_expense.html', site=site)
    
    @app.route('/site_report/<int:site_id>')
    def site_report(site_id):
        site = Site.get_summary(site_id)
        workers = SiteWorker.get_site_workers(site_id)
        materials = SiteMaterial.get_site_materials(site_id)
        expenses = SiteExpense.get_site_expenses(site_id)
        
        # Calculate totals
        total_material_cost = sum(float(m['total_cost']) for m in materials)
        total_material_paid = sum(float(m['amount_paid']) for m in materials)
        total_material_balance = sum(float(m['amount_balance']) for m in materials)
        total_expenses = sum(float(e['amount']) for e in expenses)
        
        return render_template('site_report.html',
                             site=site,
                             workers=workers,
                             materials=materials,
                             expenses=expenses,
                             total_material_cost=total_material_cost,
                             total_material_paid=total_material_paid,
                             total_material_balance=total_material_balance,
                             total_expenses=total_expenses)
    @app.route('/employee_report/<int:emp_id>')
    def employee_report(emp_id):
        """View individual employee's complete report"""
        
        # Get employee details
        employee = Employee.get_by_id(emp_id)
        if not employee:
            flash('Employee not found!', 'error')
            return redirect(url_for('employees'))
        
        # Get date range (default: current month)
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        if not start_date or not end_date:
            today = datetime.now()
            start_date = today.replace(day=1).strftime('%Y-%m-%d')
            # Last day of month
            if today.month == 12:
                end_date = today.replace(day=31).strftime('%Y-%m-%d')
            else:
                end_date = (today.replace(month=today.month + 1, day=1) - timedelta(days=1)).strftime('%Y-%m-%d')
        
        # Get attendance records
        conn = get_db()
        if not conn:
            flash('Database connection error!', 'error')
            return redirect(url_for('employees'))
        
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        # Attendance details
        cursor.execute('''
            SELECT date, status
            FROM attendance
            WHERE employee_id = %s AND date BETWEEN %s AND %s
            ORDER BY date DESC
        ''', (emp_id, start_date, end_date))
        attendance_records = cursor.fetchall()
        
        # Count present/absent days
        present_days = sum(1 for rec in attendance_records if rec['status'] == 'Present')
        absent_days = sum(1 for rec in attendance_records if rec['status'] == 'Absent')
        
        # Calculate salary
        gross_salary = present_days * float(employee['daily_salary'])
        
        # Get advances in this period
        cursor.execute('''
            SELECT date, amount, reason
            FROM advances
            WHERE employee_id = %s AND date BETWEEN %s AND %s
            ORDER BY date DESC
        ''', (emp_id, start_date, end_date))
        advances = cursor.fetchall()
        
        total_advance = sum(float(adv['amount']) for adv in advances)
        net_salary = gross_salary - total_advance
        
        # Get sites where employee is working
        cursor.execute('''
            SELECT s.id, s.site_name, s.location, sw.assigned_date, sw.role_at_site
            FROM site_workers sw
            JOIN sites s ON sw.site_id = s.id
            WHERE sw.employee_id = %s AND sw.is_active = TRUE
        ''', (emp_id,))
        current_sites = cursor.fetchall()
        
        # Calculate monthly breakdown
        cursor.execute('''
                SELECT 
                    DATE_FORMAT(date, '%%Y-%%m') AS month,
                    SUM(status = 'Present') AS present,
                    SUM(status = 'Absent') AS absent
                FROM attendance
                WHERE employee_id = %s
                GROUP BY DATE_FORMAT(date, '%%Y-%%m')
                ORDER BY month DESC
                LIMIT 6
            ''', (emp_id,))

        monthly_stats = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        return render_template('employee_report.html',
                            employee=employee,
                            start_date=start_date,
                            end_date=end_date,
                            attendance_records=attendance_records,
                            present_days=present_days,
                            absent_days=absent_days,
                            gross_salary=gross_salary,
                            advances=advances,
                            total_advance=total_advance,
                            net_salary=net_salary,
                            current_sites=current_sites,
                            monthly_stats=monthly_stats)

    @app.route('/employee_report_pdf/<int:emp_id>')
    def employee_report_pdf(emp_id):
        """Export employee report to PDF"""
        
        employee = Employee.get_by_id(emp_id)
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        # Get all data (same as above)
        conn = get_db()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        cursor.execute('''
            SELECT date, status
            FROM attendance
            WHERE employee_id = %s AND date BETWEEN %s AND %s
            ORDER BY date
        ''', (emp_id, start_date, end_date))
        attendance_records = cursor.fetchall()
        
        present_days = sum(1 for rec in attendance_records if rec['status'] == 'Present')
        gross_salary = present_days * float(employee['daily_salary'])
        
        cursor.execute('''
            SELECT date, amount, reason
            FROM advances
            WHERE employee_id = %s AND date BETWEEN %s AND %s
            ORDER BY date
        ''', (emp_id, start_date, end_date))
        advances = cursor.fetchall()
        
        total_advance = sum(float(adv['amount']) for adv in advances)
        net_salary = gross_salary - total_advance
        
        cursor.close()
        conn.close()
        
        # Create PDF
        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'],
                                    fontSize=18, spaceAfter=10, alignment=TA_CENTER)
        
        elements.append(Paragraph('SAVUNADRY CONSTRUCTION', title_style))
        elements.append(Paragraph('Employee Salary Report', styles['Heading2']))
        elements.append(Spacer(1, 20))
        
        # Employee Details
        emp_data = [
            ['Employee Name:', employee['name']],
            ['Employee ID:', str(employee['id'])],
            ['Role:', employee['role']],
            ['Daily Salary:', f"₹{employee['daily_salary']:.2f}"],
            ['Period:', f"{start_date} to {end_date}"]
        ]
        
        emp_table = Table(emp_data, colWidths=[2*inch, 4*inch])
        emp_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f0f0f0')),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(emp_table)
        elements.append(Spacer(1, 20))
        
        # Salary Summary
        summary_data = [
            ['Description', 'Amount'],
            ['Present Days', str(present_days)],
            ['Gross Salary', f"₹{gross_salary:.2f}"],
            ['Total Advance', f"₹{total_advance:.2f}"],
            ['Net Salary', f"₹{net_salary:.2f}"]
        ]
        
        summary_table = Table(summary_data, colWidths=[3*inch, 3*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E7E6E6')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(summary_table)
        
        doc.build(elements)
        output.seek(0)
        
        return send_file(output, mimetype='application/pdf', as_attachment=True,
                        download_name=f'employee_report_{employee["name"]}_{start_date}.pdf')

    @app.route('/all_employees_summary')
    def all_employees_summary():
        """Summary report of all employees"""
        
        # Get date range (default: current month)
        month = request.args.get('month', f'{datetime.now().month:02d}')
        year = request.args.get('year', str(datetime.now().year))
        
        employees = Employee.get_all()
        employee_summaries = []
        
        for emp in employees:
            present_days = Attendance.get_month_attendance(emp['id'], month, year)
            gross_salary = present_days * emp['daily_salary']
            total_advance = Advance.get_month_advance(emp['id'], month, year)
            net_salary = float(gross_salary) - float(total_advance)
            
            employee_summaries.append({
                'employee': emp,
                'present_days': present_days,
                'gross_salary': float(gross_salary),
                'total_advance': float(total_advance),
                'net_salary': net_salary
            })
        
        return render_template('all_employees_summary.html',
                            employee_summaries=employee_summaries,
                            month=month,
                            year=year)